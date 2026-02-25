import os
import time
import hashlib
from datetime import datetime, timedelta, timezone
from typing import Dict, Any, Optional, Tuple, List

import requests
from openpyxl import load_workbook

# ===== CONFIG =====
DEBUG = False  # поставь True, если снова понадобится лог
API_BASE = "https://api.clashroyale.com/v1/players/"
EXCEL_PATH = os.environ.get("MM_XLSX_PATH", "/app/data/mm.xlsx")

SHEET_ACCOUNTS = "Accounts"
SHEET_HISTORY = "CC_History"
SHEET_ARCH = "Archetypes"
SHEET_SETTINGS = "Settings"

SLEEP_BETWEEN_REQUESTS = 0.25


def encode_tag(tag: str) -> str:
    tag = tag.strip().upper()
    if not tag.startswith("#"):
        tag = "#" + tag
    return tag.replace("#", "%23")


def parse_battle_time(s: str) -> Optional[datetime]:
    # Ожидаемый формат: YYYYMMDDTHHMMSS.000Z
    if not s or "T" not in s:
        return None
    try:
        core = s[:15]  # YYYYMMDDTHHMMSS
        dt = datetime.strptime(core, "%Y%m%dT%H%M%S")
        return dt.replace(tzinfo=timezone.utc)
    except Exception:
        return None


def deck_key_from_cards(cards: List[Dict[str, Any]]) -> str:
    ids = []
    for c in cards or []:
        cid = c.get("id")
        if cid is not None:
            ids.append(int(cid))
    ids.sort()
    return "-".join(map(str, ids))


def deck_text_from_cards(cards: List[Dict[str, Any]]) -> str:
    names = []
    for c in cards or []:
        n = c.get("name")
        if n:
            names.append(str(n))
    return ", ".join(names)


def battle_result(team_crowns: int, opp_crowns: int) -> str:
    if team_crowns > opp_crowns:
        return "W"
    if team_crowns < opp_crowns:
        return "L"
    return "D"


def safe_get(d: Dict[str, Any], path: Tuple[str, ...]) -> Optional[Any]:
    cur: Any = d
    for p in path:
        if not isinstance(cur, dict) or p not in cur:
            return None
        cur = cur[p]
    return cur


def fetch_battlelog(token: str, tag: str, timeout: int = 20) -> Tuple[int, Optional[List[Dict[str, Any]]], str]:
    url = f"{API_BASE}{encode_tag(tag)}/battlelog"
    headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}

    try:
        r = requests.get(url, headers=headers, timeout=timeout)
        if r.status_code != 200:
            note = f"HTTP {r.status_code}"
            if r.status_code == 401:
                note = "401: token invalid/revoked"
            elif r.status_code == 403:
                note = "403: allowed IP mismatch in developer portal"
            elif r.status_code == 429:
                note = "429: rate limited (try less frequent)"
            return r.status_code, None, note
        data = r.json()
        if not isinstance(data, list):
            return 200, None, "unexpected JSON (not a list)"
        return 200, data, ""
    except requests.RequestException as e:
        return 0, None, f"network error: {e}"


def make_battle_id(account_tag: str, battle_time_raw: str, opp_tag: str, our_deck_key: str, mode_type: str) -> str:
    raw = f"{account_tag}|{battle_time_raw}|{opp_tag}|{our_deck_key}|{mode_type}"
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()


def load_archetypes(ws_arch) -> Dict[str, str]:
    mapping = {}
    row = 2
    while True:
        dk = ws_arch.cell(row=row, column=1).value
        if dk is None or str(dk).strip() == "":
            break
        arch = ws_arch.cell(row=row, column=2).value
        mapping[str(dk).strip()] = str(arch).strip() if arch else ""
        row += 1
    return mapping


def existing_battle_ids(ws_hist) -> set:
    ids = set()
    row = 2
    while True:
        bid = ws_hist.cell(row=row, column=1).value
        if bid is None or str(bid).strip() == "":
            break
        ids.add(str(bid).strip())
        row += 1
    return ids


def find_next_row(ws) -> int:
    r = 2
    while ws.cell(row=r, column=1).value not in (None, ""):
        r += 1
    return r


def main():
    wb = load_workbook(EXCEL_PATH)
    ws_acc = wb[SHEET_ACCOUNTS]
    ws_hist = wb[SHEET_HISTORY]
    ws_arch = wb[SHEET_ARCH]
    ws_set = wb[SHEET_SETTINGS]

    token_env = str(ws_set["B2"].value or "CLASH_TOKEN").strip()
    lookback_days = float(ws_set["B3"].value or 3)
    mode_filter = str(ws_set["B4"].value or "CC_ONLY").strip().upper()

    token = os.environ.get(token_env, "").strip()
    if len(token) < 20:
        raise SystemExit(f"Env var {token_env} is not set (or too short).")

    cutoff = datetime.now(timezone.utc) - timedelta(days=lookback_days)

    arch_map = load_archetypes(ws_arch)
    known_ids = existing_battle_ids(ws_hist)
    out_row = find_next_row(ws_hist)

    # read accounts (enabled only)
    accounts = []
    r = 2
    while True:
        t = ws_acc.cell(row=r, column=1).value
        if t is None or str(t).strip() == "":
            break
        enabled_val = ws_acc.cell(row=r, column=2).value
        if bool(enabled_val):
            accounts.append(str(t).strip().upper())
        r += 1

    if DEBUG:
        print("ModeFilter:", mode_filter, "LookbackDays:", lookback_days, "Accounts:", accounts)

    # counters
    cnt_total = 0
    cnt_old = 0
    cnt_nomode = 0
    cnt_missing_team = 0
    cnt_dupe = 0
    cnt_added = 0

    for acc_tag in accounts:
        status, battles, note = fetch_battlelog(token, acc_tag)
        if DEBUG:
            print(acc_tag, "status:", status, "battles:", (len(battles) if battles else 0), "note:", note)

        if status != 200 or not battles:
            time.sleep(SLEEP_BETWEEN_REQUESTS)
            continue

        for b in battles:
            cnt_total += 1

            bt_raw = b.get("battleTime")
            bt = parse_battle_time(bt_raw)
            if not bt or bt < cutoff:
                cnt_old += 1
                continue

            btype = str(b.get("type") or "").strip()
            gm_name = str(safe_get(b, ("gameMode", "name")) or "").strip()

            # ===== MODE FILTERING =====
            if mode_filter == "ALL":
                pass

            elif mode_filter == "CC_ONLY":
                # Подтверждённый маркер CC по твоим данным:
                # ModeType: trail
                # GameModeName: Challenge_AllCards_EventDeck_NoSet
                if not (btype == "trail" and gm_name == "Challenge_AllCards_EventDeck_NoSet"):
                    cnt_nomode += 1
                    continue

            else:
                # если захочешь фильтровать по точному type, например "trail"
                if btype.upper() != mode_filter:
                    cnt_nomode += 1
                    continue

            team = (b.get("team") or [])
            opp = (b.get("opponent") or [])
            if not team or not opp:
                cnt_missing_team += 1
                continue

            team0 = team[0]
            opp0 = opp[0]

            opp_tag = str(opp0.get("tag") or "").strip().upper()

            team_cards = team0.get("cards") or []
            opp_cards = opp0.get("cards") or []

            our_dk = deck_key_from_cards(team_cards)
            opp_dk = deck_key_from_cards(opp_cards)

            our_deck = deck_text_from_cards(team_cards)
            opp_deck = deck_text_from_cards(opp_cards)

            ctitle = str(b.get("challengeTitle") or "")

            tc = int(team0.get("crowns") or 0)
            oc = int(opp0.get("crowns") or 0)
            res = battle_result(tc, oc)

            bid = make_battle_id(acc_tag, str(bt_raw), opp_tag, our_dk, btype)
            if bid in known_ids:
                cnt_dupe += 1
                continue

            opp_arch = arch_map.get(opp_dk, "")

            row = out_row
            ws_hist.cell(row=row, column=1).value = bid
            ws_hist.cell(row=row, column=2).value = acc_tag
            ws_hist.cell(row=row, column=3).value = bt.strftime("%Y-%m-%d %H:%M:%S")
            ws_hist.cell(row=row, column=4).value = btype
            ws_hist.cell(row=row, column=5).value = gm_name
            ws_hist.cell(row=row, column=6).value = ctitle
            ws_hist.cell(row=row, column=7).value = opp_tag
            ws_hist.cell(row=row, column=8).value = res
            ws_hist.cell(row=row, column=9).value = our_dk
            ws_hist.cell(row=row, column=10).value = our_deck
            ws_hist.cell(row=row, column=11).value = opp_dk
            ws_hist.cell(row=row, column=12).value = opp_deck
            ws_hist.cell(row=row, column=13).value = opp_arch
            # Notes (14) оставляем пустым — ты заполняешь вручную

            out_row += 1
            known_ids.add(bid)
            cnt_added += 1

        time.sleep(SLEEP_BETWEEN_REQUESTS)

    wb.save(EXCEL_PATH)

    print(f"Done. Added {cnt_added} new matches into {EXCEL_PATH}.")
    if DEBUG:
        print("Scanned:", cnt_total, "| old:", cnt_old, "| mode-skip:", cnt_nomode,
              "| missing team:", cnt_missing_team, "| dupes:", cnt_dupe, "| added:", cnt_added)


if __name__ == "__main__":
    main()
